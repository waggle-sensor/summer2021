import statistics
import cv2
import numpy as np
import math
import pandas as pd
from pathlib import Path
import os
from openpyxl import load_workbook

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """

    Helper method taken from:
    https://stackoverflow.com/questions/38074678/append-existing-excel-sheet-with-new-dataframe-using-python-pandas/38075046#38075046

    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def display(image_to_display):
    '''
    Helper function that displays an image passed in
    :param image_to_display: image that needs to be displayed
    :return: void
    '''
    cv2.imshow("image", image_to_display)
    cv2.waitKey(0)
    cv2.destroyAllWindows()

def find_mean_hue(hsv_image):
    '''
    Convert hue angles to a set of vectors from polar to cartesian coordinates
    after taking mean of those coordinates, convert back to polar form
    :param hsv_image: image in hsv format
    :return: mean hue
    '''

    hsv_1D = hsv_image[...,0].flatten()

    x_list = []
    y_list = []

    for i in range(len(hsv_1D)):

        hsv_1D[i] *= 2

        angle_x = math.radians(hsv_1D[i])
        angle_y = math.radians(hsv_1D[i])

        x = math.cos(angle_x)
        y = math.sin(angle_y)

        x_list.append(x)
        y_list.append(y)

    x_mean = statistics.mean(x_list)
    y_mean = statistics.mean(y_list)

    mean_angle = ((math.atan(y_mean / x_mean)) * 180) / math.pi

    if ((x_mean < 0 and y_mean < 0) or (x_mean < 0 and y_mean > 0)):
        mean_angle += 180
    elif (x_mean > 0 and y_mean < 0):
        mean_angle += 360

    mean_hue = mean_angle / 2

    return mean_hue

def find_mean_hsv(hsv_image):
    '''
    Computes the mean hue, saturation, and value of image given HSV image
    :param hsv_image: image in hsv format
    :return: mean hue, saturation, and value in list
    '''

    mean_hue = find_mean_hue(hsv_image)
    mean_sat = hsv_image[..., 1].mean()
    mean_val = hsv_image[...,2].mean()

    mean_hsv = [mean_hue, mean_sat, mean_val]

    return mean_hsv

def find_standard_deviation_hsv(hsv_image):
    '''
    computed standard deviation of hsv in image
    :param hsv_image: image in hsv format
    :return: array of hsv standard deviation
    '''

    hue_1D = hsv_image[...,0].flatten()
    sat_1D = hsv_image[...,1].flatten()
    val_1D = hsv_image[...,2].flatten()

    hue_SD = np.std(hue_1D)
    sat_SD = np.std(sat_1D)
    val_SD = np.std(val_1D)

    SD_hsv = [hue_SD, sat_SD, val_SD]

    return SD_hsv

def detect_edges(image, lower_threshold, upper_threshold):
    '''
    Use canny edge detection to detect edges in an image
    :param image: image whose edges are to be detected
    :param lower_threshold: lower threshold for canny edge detector
    :param upper_threshold: upper threshold for canny edge detector
    :return: image with edges detected
    '''

    image = cv2.bilateralFilter(image, 9, 75, 75)
    canny = cv2.Canny(image, lower_threshold, upper_threshold)

    return canny

def find_straight_edge_density(image):
  '''
  Find a measure of how many straight edges are in the scene image
  using fast line detector method. Blur image while keeping edges
  sharp using bilateral filter to reduce noise
  :return: sum of straight lines
  '''

  gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

  gray_image = cv2.bilateralFilter(gray_image, 9, 75, 75)

  fld = cv2.ximgproc.createFastLineDetector()
  lines = fld.detect(gray_image)
  draw_image = fld.drawSegments(image, lines) # can use to display image

  try:
    if len(lines) == 0:
        return 0
  except:
      return 0

  sum_length = 0
  for i in range(len(lines)):
      x1 = lines[i][0][0]
      y1 = lines[i][0][1]
      x2 = lines[i][0][2]
      y2 = lines[i][0][3]

      length_line_segment = math.sqrt(((x2 - x1) ** 2) +  ((y2 - y1) ** 2))

      sum_length += length_line_segment

  return sum_length


def find_edge_density(image):
    '''
    Computes how much of the image consists of edges
    :param image: BGR image whose edge density needs to be find
    :return: percentage of image that is made up of edges
    '''

    edges_image = detect_edges(image, 100, 200)

    total_pixels = edges_image.shape[0] * edges_image.shape[1]
    white_pixels = cv2.countNonZero(edges_image)

    edge_density_percentage = (white_pixels / total_pixels) * 100

    return edge_density_percentage

def find_entropy(image):
    '''
    The entropy or average information of an image is a measure of the degree of randomness in the image.
    :param image: BGR image
    :return: entropy
    '''

    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    image_1D = gray_image.flatten()

    length = image_1D.size
    symset = list(set(image_1D))
    propab = [np.size(image_1D[image_1D == i]) / (1.0 * length) for i in symset]
    entropy = np.sum([p * np.log2(1.0/p) for p in propab])

    return entropy

def create_base_dataframe():
    '''
    creates base dataframe to store data
    :return: dataframe
    '''

    df = pd.read_excel(r'MIT_ImageRatings.xlsx')

    df = df.rename(columns={"Hue": "Mean Hue"})
    df = df.rename(columns={"Sat": "Mean Sat"})
    df = df.rename(columns={"Lum": "Mean Value"})
    df = df.rename(columns={"sdBright": "sdValue"})
    del df['LRSymm']
    del df['UDSymm']
    del df['NSED']

    return df

def add_features():
    '''
     Computes features of images that would help measure visual disorder
    :return: dataframe with computed features for each image in training data
    '''

    data_file_path = r'augmented_data.xlsx'
    df = pd.read_excel(data_file_path)

    rows_to_delete = []

    start_row = 1105
    for i in range(start_row, len(df)):

        image_name = df.loc[i].at["originalName"]
        folder = "training_images/TFK/"
        file_path = folder + image_name

        my_file = Path(file_path)
        if (my_file.is_file()):
            image = cv2.imread(file_path)
            hsv_image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

            mean_hsv = find_mean_hsv(hsv_image)
            standard_dev_hsv = find_standard_deviation_hsv(hsv_image)
            edge_density = find_edge_density(image)
            straight_edge_density = find_straight_edge_density(image)
            image = cv2.imread(file_path)
            entropy = find_entropy(image)

            # replace values with the features that I computed
            df.loc[i, 'Mean Hue'] = mean_hsv[0]
            df.loc[i, 'Mean Sat'] = mean_hsv[1]
            df.loc[i, 'Mean Value'] = mean_hsv[2]
            df.loc[i, 'sdHue'] = standard_dev_hsv[0]
            df.loc[i, 'sdSat'] = standard_dev_hsv[1]
            df.loc[i, 'sdValue'] = standard_dev_hsv[2]
            df.loc[i, 'ED'] = edge_density
            df.loc[i, 'SED'] = straight_edge_density
            df.loc[i, 'Entropy'] = entropy
        else:
            rows_to_delete.append(i)

        append_df_to_excel(r'normalized_data.xlsx', df, startrow= 0)

    # delete all the rows that did not have images associated with them
    for i in range(len(rows_to_delete)):
        df.drop(labels=rows_to_delete[i], axis = 0, inplace = True)

    append_df_to_excel(r'normalized_data.xlsx', df, startrow=0)

    return df
