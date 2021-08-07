import cv2
import pandas as pd
from pathlib import Path
import os
from openpyxl import load_workbook
from detect_features import find_mean_hsv
from detect_features import find_standard_deviation_hsv
from detect_features import find_edge_density
from detect_features import find_straight_edge_density
from detect_features import find_entropy
from data_augment import brightness
from data_augment import change_hsv
from data_augment import horizontal_shift
from data_augment import vertical_shift
from data_augment import random_noise
from data_augment import zoom


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """

    Helper method taken from:
    https://stackoverflow.com/questions/38074678/append-existing-excel-sheet-with-new-dataframe-using-python-pandas/38075046#38075046

    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def add_augmented_images_to_dataframe(file_with_image_names, file_to_add_data):
    '''
     Computes features of images that would help measure visual disorder
    :return: dataframe with computed features for each image in training data
    '''

    df = pd.read_excel(file_with_image_names)

    for i in range(len(df)):

        image_name = df.loc[i].at["originalName"]
        folder = "training_images/MIT_images/"
        file_path = folder + image_name

        my_file = Path(file_path)
        if (my_file.is_file()):
            image = cv2.imread(file_path)

            # pick data augmentations
            image = change_hsv(image)
            image = brightness(image, 1.2)
            image = horizontal_shift(image, 0.5)

            hsv_image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

            mean_hsv = find_mean_hsv(hsv_image)
            standard_dev_hsv = find_standard_deviation_hsv(hsv_image)
            edge_density = find_edge_density(image)
            straight_edge_density = find_straight_edge_density(image)
            image = cv2.imread(file_path)
            entropy = find_entropy(image)

            # replace values with the features that I computed
            df.loc[i, 'Mean Hue'] = mean_hsv[0]
            df.loc[i, 'Mean Sat'] = mean_hsv[1]
            df.loc[i, 'Mean Value'] = mean_hsv[2]
            df.loc[i, 'sdHue'] = standard_dev_hsv[0]
            df.loc[i, 'sdSat'] = standard_dev_hsv[1]
            df.loc[i, 'sdValue'] = standard_dev_hsv[2]
            df.loc[i, 'ED'] = edge_density
            df.loc[i, 'SED'] = straight_edge_density
            df.loc[i, 'Entropy'] = entropy

            append_df_to_excel(file_to_add_data,df,startrow=0)

def add_features(file_with_image_names, file_to_add_data):
    '''
     Computes features of images that would help measure visual disorder
    :return: dataframe with computed features for each image in training data
    '''

    df = pd.read_excel(file_with_image_names)

    rows_to_delete = []

    start_row = 1105
    for i in range(start_row, len(df)):

        image_name = df.loc[i].at["originalName"]
        folder = "training_images/TFK/"
        file_path = folder + image_name

        my_file = Path(file_path)
        print(file_path)
        image = cv2.imread(file_path)
        hsv_image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

        mean_hsv = find_mean_hsv(hsv_image)
        standard_dev_hsv = find_standard_deviation_hsv(hsv_image)
        edge_density = find_edge_density(image)
        straight_edge_density = find_straight_edge_density(image)
        image = cv2.imread(file_path)
        entropy = find_entropy(image)

        # replace values with the features that I computed
        df.loc[i, 'Mean Hue'] = mean_hsv[0]
        df.loc[i, 'Mean Sat'] = mean_hsv[1]
        df.loc[i, 'Mean Value'] = mean_hsv[2]
        df.loc[i, 'sdHue'] = standard_dev_hsv[0]
        df.loc[i, 'sdSat'] = standard_dev_hsv[1]
        df.loc[i, 'sdValue'] = standard_dev_hsv[2]
        df.loc[i, 'ED'] = edge_density
        df.loc[i, 'SED'] = straight_edge_density
        df.loc[i, 'Entropy'] = entropy

        append_df_to_excel(file_to_add_data, df, startrow=0)

    # delete all the rows that did not have images associated with them
    for i in range(len(rows_to_delete)):
        df.drop(labels=rows_to_delete[i], axis=0, inplace=True)

    append_df_to_excel(r'dataframes/normalized_data.xlsx', df, startrow=0)

    return df

def main():
    file_with_image_names = r'example.xlsx'
    file_to_add_data = r'example.xlsx'
    add_features(file_with_image_names, file_to_add_data)

if __name__ == '__main__':
    main()